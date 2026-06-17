from io import BytesIO

from openpyxl import Workbook

from app.rag.parser import detect_file_type, parse_bytes


def _make_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "产品表"
    ws.append(["产品名", "价格", "库存"])
    ws.append(["货币基金", 1000, 50])
    ws.append(["债券基金", 2000, 30])
    ws2 = wb.create_sheet("说明")
    ws2.append(["备注"])
    ws2.append(["本表为示例"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_extracts_sheet_and_cells():
    data = _make_xlsx()
    text = parse_bytes("test.xlsx", data)
    assert "# 产品表" in text
    assert "产品名\t价格\t库存" in text
    assert "货币基金\t1000\t50" in text
    assert "# 说明" in text
    assert "本表为示例" in text


def test_detect_file_type_xlsx():
    assert detect_file_type("a.xlsx") == "xlsx"
    assert detect_file_type("b.xls") == "xlsx"
    assert detect_file_type("c.pdf") == "pdf"
    assert detect_file_type("d.docx") == "docx"
    assert detect_file_type("e.md") == "md"
    assert detect_file_type("f.txt") == "txt"
    assert detect_file_type("g.unknown") == "other"


def test_parse_bytes_txt_fallback():
    text = parse_bytes("note.txt", "你好 world".encode("utf-8"))
    assert text == "你好 world"


def test_parse_bytes_empty_sheet_skipped():
    wb = Workbook()
    wb.active.title = "Empty"
    buf = BytesIO()
    wb.save(buf)
    text = parse_bytes("empty.xlsx", buf.getvalue())
    # 空表应被跳过，返回空字符串
    assert text == ""

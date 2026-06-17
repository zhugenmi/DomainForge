from __future__ import annotations

from io import BytesIO


def parse_xlsx(data: bytes | None = None, path: str | None = None) -> str:
    """从 Excel 文件抽取文本。每个 sheet 输出标题行 + 制表符拼接的单元格行。

    使用 read_only + data_only 模式：避免公式计算，流式读取降低内存。
    空 sheet 跳过；合并单元格只在左上角有值（openpyxl 默认行为）。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""

    if path is not None:
        wb = load_workbook(path, read_only=True, data_only=True)
    elif data is not None:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    else:
        return ""

    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines: list[str] = [f"# {sheet_name}"]
            non_empty_rows = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = "\t".join(cells).strip()
                if line:
                    sheet_lines.append(line)
                    non_empty_rows += 1
            if non_empty_rows > 0:
                parts.append("\n".join(sheet_lines))
    finally:
        wb.close()
    return "\n\n".join(parts).strip()


__all__ = ["parse_xlsx"]
